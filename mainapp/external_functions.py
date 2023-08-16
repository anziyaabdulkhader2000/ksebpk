import smtplib
import ssl
from email.message import EmailMessage

import qrcode
import random
from datetime import datetime
from dateutil.relativedelta import relativedelta


#________________________ email sender __________________________________________________

def email_sender(To):
    sender = "aswindas.209official@gmail.com"
    password = "zutmiuhnzbauohkd"
    receiver = To

    subject = "Pension Management System Account Set-up"

    # HTML body with an embedded image and form
    body = f"""
    <html>
    
        <body>
            <p>Your Account is Varified by Admin</p>
            <p>Scan this qr code to get your Transaction ID and Subscribe</p>
            <p style="color:red"><strong>Do not share this with others.</strong></p>
            <p><img src="cid:image1"></p>
            <p>Login to your account with email id and password to complete the transaction</p>
            <p style="color:rgb(62, 62, 232)"><strong>Your account will be activated once you subscribe</strong></p>

        </body>
    </html>
    """

    em = EmailMessage()
    em['From'] = sender
    em['To'] = receiver
    em['Subject'] = subject
    em.add_alternative(body, subtype='html')

    # Attach the image to the email
    with open("qrcode.png", "rb") as image_file:
        image_data = image_file.read()
        em.get_payload()[0].add_related(image_data, 'image', 'png', cid='image1')

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(sender, password)
        smtp.sendmail(sender, receiver, em.as_string())





def generate_qr_code(data, filename):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    img.save(filename)
    print(f"QR code generated and saved as '{filename}'")


def qr_code_generator_main(data_to_encode):
    # Replace 'Your Data Here' with the data you want to encode in the QR code

    output_file = "qrcode.png"

    generate_qr_code(data_to_encode, output_file)

# Call the QR code generator function



def generate_transaction_id():
    # Get the current date in the format 'YYYYMMDD'
    current_date = datetime.now().strftime("%Y%m%d")

    # Generate a random 6-digit number
    random_number = random.randint(10**5, (10**6)-1)

    # Combine the current date and the random number to form a 12-digit transaction ID
    transaction_id = int(f"{current_date}{random_number}")

    return transaction_id

# Example usage


def qr_sender_main(receiver_mail):
    transaction_id = generate_transaction_id()
    qr_code_generator_main(transaction_id)
    email_sender(receiver_mail)
    return transaction_id

#________________________XL sheet creater__________________________________________________

def xl_namer(file_path):
    today = datetime.today()

    year_str = str(today.year)
    month_str = str(today.month).zfill(2)  # Zero-padding month with leading zeros if necessary
    day_str = str(today.day).zfill(2)  # Zero-padding day with leading zeros if necessary
    hour_str = str(today.hour).zfill(2)  # Zero-padding hour with leading zeros if necessary
    minute_str = str(today.minute).zfill(2)  # Zero-padding minute with leading zeros if necessary
    second_str = str(today.second).zfill(2) #

    # Join the com{ponents with desired separators
    file_name = f"{file_path}\Subscription_data-{year_str}-{month_str}-{day_str}-{hour_str}-{minute_str}-{second_str}.xlsx"

    return file_name




def xl_creater(title,data_pool,file_path):
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        data_pool = data_pool
        data = []

        for user in data_pool:
            each = {'User_name': user['User_name'], 'User_email': user["User_email"], 'User_ARU': user["User_ARU"], 'User_phone': user["User_phone"]}
            data.append(each)

        # Create an empty DataFrame
        df = pd.DataFrame(data)

        # Create a new Excel workbook
        wb = Workbook()

        # Select the default active sheet
        ws = wb.active

        # Add the main title at the top of the sheet and merge cells
        title_cell = ws.cell(row=2, column=1, value=title)
        title_cell.font = Font(size=18, bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=len(df.columns)+5)

        # Add an empty row after the main title
        ws.append([])

        # Get the keys from the DataFrame to use as column headers
        column_headers = list(df.columns)

        # Rearrange the columns in the desired order (Name, Email, Area, Phone)
        desired_order = ['User_name', 'User_email', 'User_ARU', 'User_phone']
        column_headers = [col for col in desired_order if col in df.columns]

        # Add the column headers and adjust column width
        for col_idx, column_name in enumerate(column_headers, start=1):
            column_letter = get_column_letter(col_idx)
            col_title = ws.cell(row=6, column=col_idx, value=column_name)
            col_title.font = Font(size=12, bold=True)
            ws.column_dimensions[column_letter].width = max(len(str(column_name)), len(str(df[column_name].max()))) + 2

        # Add the values from the DataFrame as rows and adjust row height
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                column_letter = get_column_letter(col_idx)
                ws.cell(row=row_idx + 7, column=col_idx, value=value)
                cell_value_length = len(str(value))
                if ws.column_dimensions[column_letter].width < cell_value_length + 2:
                    ws.column_dimensions[column_letter].width = cell_value_length + 2
            ws.row_dimensions[row_idx + 4].height = 14

        file_name = xl_namer(file_path)
        wb.save(file_name)

        return True   # returning the file

    except:
        return False
